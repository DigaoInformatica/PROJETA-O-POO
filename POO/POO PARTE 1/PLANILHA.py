import openpyxl

# Criar uma planilha 
trabalho = openpyxl.Workbook()

# Remover pagina 'Sheet'
default_sheet = trabalho['Sheet']
trabalho.remove(default_sheet)

# Criar uma página 'LIVROS'
trabalho.create_sheet('LIVROS')

# Criar uma página 'USUARIOS'
trabalho.create_sheet('USUARIOS')

# Criar uma página 'RESERVAS'
trabalho.create_sheet('RESERVAS')

# Salvar o arquivo no arquivo 'POO.xlsx'
trabalho.save('POO.xlsx')