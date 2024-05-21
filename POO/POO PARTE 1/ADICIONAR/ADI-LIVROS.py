import openpyxl

# Carregar o arquivo existente
trabalho = openpyxl.load_workbook('POO.xlsx')

# Selecionar a aba 'LIVROS'
livro_page = trabalho['LIVROS']

# Função para adicionar dados com input
def adicionar_livro():
    dados = [
        input("Digite o título do livro: "),
        input("Digite o autor do livro: "),
        input("Digite o gênero do livro: "),
        input("Digite o tipo do livro: "),
        input("Digite o status do livro (Disponível/Indisponível): ")
    ]

    linha = 6
    while livro_page.cell(row=linha, column=1).value is not None:
        linha += 1

    for col, valor in enumerate(dados, start=2):
        livro_page.cell(row=linha, column=col, value=valor)

# Loop para adicionar vários livros
while True:
    adicionar_livro()
    if input("Deseja adicionar outro livro? (s/n): ").strip().lower() != 's':
        break

# Salvar o arquivo
trabalho.save('POO.xlsx')
print("Dados salvos com sucesso.")