import openpyxl

# Carregar o arquivo existente
trabalho = openpyxl.load_workbook('POO.xlsx')

# Selecionar a aba 'USUARIOS'
usuarios_page = trabalho['USUARIOS']

# Função para adicionar dados com input
def adicionar_usuario():
    dados = [
        input("Digite a ID do usuário: "),
        input("Digite o nome do usuário: "),
        input("Digite o CPF do usuário: "),
        input("Digite o tipo do usuário (Funcionário/Usuário): "),
        input("Digite a senha do usuário: ")
    ]

    linha = 6
    while usuarios_page.cell(row=linha, column=1).value is not None:
        linha += 1

    for col, valor in enumerate(dados, start=2):
        usuarios_page.cell(row=linha, columRn=col, value=valor)

# Loop para adicionar vários usuários
while True:
    adicionar_usuario()
    if input("Deseja adicionar outro usuário? (s/n): ").strip().lower() != 's':
        break

# Salvar o arquivo
trabalho.save('POO.xlsx')
print("Dados salvos com sucesso.")
