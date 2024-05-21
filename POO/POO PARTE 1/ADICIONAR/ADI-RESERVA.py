import openpyxl

# Carregar o arquivo existente
trabalho = openpyxl.load_workbook('POO.xlsx')

# Selecionar a aba 'RESERVAS'
reservas_page = trabalho['RESERVAS']

# Função para adicionar dados com input
def adicionar_reserva():
    dados = [
        input("Digite o nome do cliente: "),
        input("Digite o livro reservado: "),
        input("Digite a data da reserva: "),
        input("Digite a data de retirada: "),
        input("Digite o status da reserva (Ativa/Cancelada): ")
    ]

    linha = 3
    while reservas_page.cell(row=linha, column=1).value is not None:
        linha += 1

    for col, valor in enumerate(dados, start=1):
        reservas_page.cell(row=linha, column=col, value=valor)

# Loop para adicionar várias reservas
while True:
    adicionar_reserva()
    if input("Deseja adicionar outra reserva? (s/n): ").strip().lower() != 's':
        break

# Salvar o arquivo
trabalho.save('POO.xlsx')
print("Dados salvos com sucesso.")
