import openpyxl

# Carregar a planilha
nome_arquivo = 'POO.xlsx'
nome_aba = 'USUARIOS'
planilha = openpyxl.load_workbook(nome_arquivo)[nome_aba]

# Inicializar a contagem
contador = 0

# Listar Usuarios e suas informações
print("Usuários disponíveis e informações:")
for row in planilha.iter_rows(min_row=6, max_row=planilha.max_row, min_col=2, max_col=6):
    if row[0].value:  # Verificar se há um título na célula
        # Incrementar a contagem
        contador += 1
        # Exibir as informações do título, autor, gênero, tipo e status em uma linha formatada
        print(f"{contador}. ID: {row[0].value} | Nome: {row[1].value} | CPF: {row[2].value} | Tipo: {row[3].value} | Senha: {row[4].value}")

# Remover usuario por número
numero_para_remover = int(input("Digite o número da usuário que deseja remover: "))
usuario = None
contador = 0
for row in planilha.iter_rows(min_row=6, max_row=planilha.max_row, min_col=2, max_col=6):
    if row[0].value:  # Verificar se há um título na célula
        # Incrementar a contagem
        contador += 1
        if contador == numero_para_remover:  # Verificar se o número corresponde ao usuario
            usuario = row
            break

if usuario:  # Se um usuario foi encontrado
    titulo = usuario[0].value
    print("\nInformações do usuário:")
    # Exibir as informações do usuario em uma linha formatada
    print(f"ID: {usuario[0].value} | Nome: {usuario[1].value} | CPF: {usuario[2].value} | Tipo: {usuario[3].value} | Senha: {usuario[4].value}")
    confirmacao = input("Deseja remover este usuário? (s/n): ").strip().lower()
    if confirmacao == 's':  # Se o usuário confirmar a remoção
        for cell in usuario:
            cell.value = None  # Definir o valor da célula como None
        print(f'O conteúdo do usuario "{titulo}" foi removido com sucesso.')
    else:  # Se o usuário cancelar a remoção
        print("Remoção cancelada.")
else:  # Se nenhum usuario correspondente foi encontrado
    print(f'\nNão foi possível encontrar um usuário correspondente ao número {numero_para_remover}.')

# Salvar a planilha
planilha.parent.save(nome_arquivo)