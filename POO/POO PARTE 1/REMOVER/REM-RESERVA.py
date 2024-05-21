import openpyxl

# Carregar a planilha
nome_arquivo = 'POO.xlsx'
nome_aba = 'RESERVAS'
planilha = openpyxl.load_workbook(nome_arquivo)[nome_aba]

# Inicializar a contagem
contador = 0

# Listar Reservas e suas informações
print("Reservas disponíveis e informações:")
for row in planilha.iter_rows(min_row=6, max_row=planilha.max_row, min_col=2, max_col=6):
    if row[0].value:  # Verificar se há um título na célula
        # Incrementar a contagem
        contador += 1
        # Exibir as informações do título, autor, gênero, tipo e status em uma linha formatada
        print(f"{contador}. Título: {row[0].value} | Tipo: {row[1].value} | ID Usuário: {row[2].value} | Data Reserva: {row[3].value} | Data Devolução: {row[4].value}")

# Remover reserva por número
numero_para_remover = int(input("Digite o número da reserva que deseja remover: "))
reserva = None
contador = 0
for row in planilha.iter_rows(min_row=6, max_row=planilha.max_row, min_col=2, max_col=6):
    if row[0].value:  # Verificar se há um título na célula
        # Incrementar a contagem
        contador += 1
        if contador == numero_para_remover:  # Verificar se o número corresponde ao reserva
            reserva = row
            break

if reserva:  # Se um reserva foi encontrado
    titulo = reserva[0].value
    print("\nInformações do reserva:")
    # Exibir as informações do reserva em uma linha formatada
    print(f"Título: {reserva[0].value} | Tipo: {reserva[1].value} | ID Usuário: {reserva[2].value} | Data Reserva: {reserva[3].value} | Data Devolução: {reserva[4].value}")
    confirmacao = input("Deseja remover este reserva? (s/n): ").strip().lower()
    if confirmacao == 's':  # Se o usuário confirmar a remoção
        for cell in reserva:
            cell.value = None  # Definir o valor da célula como None
        print(f'O conteúdo do reserva "{titulo}" foi removido com sucesso.')
    else:  # Se o usuário cancelar a remoção
        print("Remoção cancelada.")
else:  # Se nenhum reserva correspondente foi encontrado
    print(f'\nNão foi possível encontrar um reserva correspondente ao número {numero_para_remover}.')

# Salvar a planilha
planilha.parent.save(nome_arquivo)