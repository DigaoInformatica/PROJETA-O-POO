import openpyxl

# Carregar a planilha
nome_arquivo = 'POO.xlsx'
nome_aba = 'LIVROS'
planilha = openpyxl.load_workbook(nome_arquivo)[nome_aba]

# Inicializar a contagem
contador = 0

# Listar títulos e suas informações
print("Títulos disponíveis e suas informações:")
for row in planilha.iter_rows(min_row=6, max_row=planilha.max_row, min_col=2, max_col=6):
    if row[0].value:  # Verificar se há um título na célula
        # Incrementar a contagem
        contador += 1
        # Exibir as informações do título, autor, gênero, tipo e status em uma linha formatada
        print(f"{contador}. Título: {row[0].value} | Autor: {row[1].value} | Gênero: {row[2].value} | Tipo: {row[3].value} | Status: {row[4].value}")

# Remover livro por número
numero_para_remover = int(input("Digite o número do livro que deseja remover: "))
livro = None
contador = 0
for row in planilha.iter_rows(min_row=6, max_row=planilha.max_row, min_col=2, max_col=6):
    if row[0].value:  # Verificar se há um título na célula
        # Incrementar a contagem
        contador += 1
        if contador == numero_para_remover:  # Verificar se o número corresponde ao livro
            livro = row
            break

if livro:  # Se um livro foi encontrado
    titulo = livro[0].value
    print("\nInformações do livro:")
    # Exibir as informações do livro em uma linha formatada
    print(f"Título: {livro[0].value} | Autor: {livro[1].value} | Gênero: {livro[2].value} | Tipo: {livro[3].value} | Status: {livro[4].value}")
    confirmacao = input("Deseja remover este livro? (s/n): ").strip().lower()
    if confirmacao == 's':  # Se o usuário confirmar a remoção
        for cell in livro:
            cell.value = None  # Definir o valor da célula como None
        print(f'O conteúdo do livro "{titulo}" foi removido com sucesso.')
    else:  # Se o usuário cancelar a remoção
        print("Remoção cancelada.")
else:  # Se nenhum livro correspondente foi encontrado
    print(f'\nNão foi possível encontrar um livro correspondente ao número {numero_para_remover}.')

# Salvar a planilha
planilha.parent.save(nome_arquivo)